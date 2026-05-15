#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
회사 교육용 크로스워드 퍼즐 생성기
Dark Theme / CustomTkinter GUI
"""

import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
import random
import json
import os
import time
import copy
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import io
import sys

# ── PyInstaller 단일 파일 exe 지원용 리소스 경로 함수 ──────────────
def resource_path(relative_path: str) -> str:
    """개발 환경과 PyInstaller --onefile 환경 모두에서 올바른 절대경로를 반환한다."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, relative_path)

# ── 한글 폰트 등록 ────────────────────────────────────────────────
def _register_korean_fonts():
    """나눔고딕 TTF를 ReportLab에 등록 (한글 PDF 출력용)"""
    candidates = [
        # 나눔고딕 (Ubuntu fonts-nanum 패키지)
        ("/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
         "/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf"),
        # macOS
        ("/Library/Fonts/NanumGothic.ttf", None),
        # Windows
        ("C:/Windows/Fonts/malgun.ttf", "C:/Windows/Fonts/malgunbd.ttf"),
    ]
    for regular, bold in candidates:
        if os.path.exists(regular):
            try:
                pdfmetrics.registerFont(TTFont("KorFont",     regular))
                pdfmetrics.registerFont(TTFont("KorFont-Bold", bold if bold and os.path.exists(bold) else regular))
                from reportlab.lib.fonts import addMapping
                addMapping("KorFont", 0, 0, "KorFont")
                addMapping("KorFont", 1, 0, "KorFont-Bold")
                return True
            except Exception:
                pass
    return False

_KOR_FONT_OK = _register_korean_fonts()
KOR_FONT      = "KorFont"      if _KOR_FONT_OK else "Helvetica"
KOR_FONT_BOLD = "KorFont-Bold" if _KOR_FONT_OK else "Helvetica-Bold"

# ── Dark Theme 설정 ──────────────────────────────────────────────
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ── 색상 팔레트 ───────────────────────────────────────────────────
COLORS = {
    "bg":        "#1a1a2e",
    "panel":     "#16213e",
    "card":      "#0f3460",
    "accent":    "#e94560",
    "accent2":   "#533483",
    "gold":      "#f5a623",
    "correct":   "#2ecc71",
    "wrong":     "#e74c3c",
    "cell_bg":   "#1e2a45",
    "cell_text": "#e0e0e0",
    "cell_num":  "#a0c4ff",
    "border":    "#2d4a7a",
    "text":      "#e0e0e0",
    "subtext":   "#8899aa",
    "across":    "#1a4a7a",
    "down":      "#4a1a7a",
    "selected":  "#e94560",
    "highlight": "#2d6a9f",
    "empty":     "#0d1b2a",
}

# ════════════════════════════════════════════════════════════════
# 크로스워드 생성 엔진
# ════════════════════════════════════════════════════════════════
class CrosswordGenerator:
    def __init__(self, words_clues, grid_rows, grid_cols, max_attempts=50):
        self.words_clues = [(w.strip().upper(), c.strip()) for w, c in words_clues if 2 <= len(w.strip()) <= 15]
        self.rows = grid_rows
        self.cols = grid_cols
        self.max_attempts = max_attempts
        self.grid = []
        self.placed = []   # {word, clue, row, col, direction, number}

    def _empty_grid(self):
        return [['#'] * self.cols for _ in range(self.rows)]

    def _can_place(self, grid, word, r, c, direction):
        L = len(word)
        # ── 기본 경계 검사 (음수 좌표도 차단 → wrap-around 버그 방지) ──
        if r < 0 or c < 0: return False
        if direction == 'A':
            if c + L > self.cols: return False
            if c > 0 and grid[r][c-1] != '#': return False
            if c + L < self.cols and grid[r][c+L] != '#': return False
            for i, ch in enumerate(word):
                cell = grid[r][c+i]
                if cell != '#' and cell != ch: return False
                # 교차점 이외엔 같은 방향 인접 셀이 있으면 안 됨
                if cell == '#':
                    # 위 행이 존재하고 해당 셀이 비어있지 않으면 배치 불가
                    if r > 0 and grid[r-1][c+i] != '#': return False
                    # 아래 행이 존재하고 해당 셀이 비어있지 않으면 배치 불가
                    if r < self.rows - 1 and grid[r+1][c+i] != '#': return False
        else:  # 'D'
            if r + L > self.rows: return False
            if r > 0 and grid[r-1][c] != '#': return False
            if r + L < self.rows and grid[r+L][c] != '#': return False
            for i, ch in enumerate(word):
                cell = grid[r+i][c]
                if cell != '#' and cell != ch: return False
                if cell == '#':
                    # 왼쪽 열이 존재하고 해당 셀이 비어있지 않으면 배치 불가
                    if c > 0 and grid[r+i][c-1] != '#': return False
                    # 오른쪽 열이 존재하고 해당 셀이 비어있지 않으면 배치 불가
                    if c < self.cols - 1 and grid[r+i][c+1] != '#': return False
        return True

    def _place_word(self, grid, word, r, c, direction):
        new_grid = copy.deepcopy(grid)
        if direction == 'A':
            for i, ch in enumerate(word):
                new_grid[r][c+i] = ch
        else:
            for i, ch in enumerate(word):
                new_grid[r+i][c] = ch
        return new_grid

    def _find_positions(self, grid, word, placed_count):
        positions = []
        # 첫 단어: 중앙 배치
        if placed_count == 0:
            r = self.rows // 2
            c = (self.cols - len(word)) // 2
            if self._can_place(grid, word, r, c, 'A'):
                positions.append((r, c, 'A', 10))
            r = (self.rows - len(word)) // 2
            c = self.cols // 2
            if r >= 0 and self._can_place(grid, word, r, c, 'D'):
                positions.append((r, c, 'D', 10))
            return positions

        # 교차점 찾기
        for r in range(self.rows):
            for c in range(self.cols):
                if grid[r][c] == '#': continue
                ch = grid[r][c]
                for i, wch in enumerate(word):
                    if wch == ch:
                        # Across: 시작 열이 0 이상인지 확인
                        tr, tc = r, c - i
                        if tc >= 0 and self._can_place(grid, word, tr, tc, 'A'):
                            positions.append((tr, tc, 'A', 5))
                        # Down: 시작 행이 0 이상인지 확인
                        tr, tc = r - i, c
                        if tr >= 0 and self._can_place(grid, word, tr, tc, 'D'):
                            positions.append((tr, tc, 'D', 5))
        return positions

    def generate(self):
        best_grid, best_placed = None, []
        wc_list = list(self.words_clues)

        for attempt in range(self.max_attempts):
            random.shuffle(wc_list)
            grid = self._empty_grid()
            placed = []

            for word, clue in wc_list:
                positions = self._find_positions(grid, word, len(placed))
                if not positions: continue
                random.shuffle(positions)
                placed_ok = False
                for r, c, direction, _ in positions:
                    # ── 중복 배치 차단 ──────────────────────────────
                    # 1) 동일 단어가 이미 배치된 경우
                    if any(p['word'] == word for p in placed):
                        continue
                    # 2) 같은 방향·같은 시작점에 다른 단어가 이미 있는 경우
                    #    (ES 배치 후 ESS가 같은 위치 같은 방향으로 들어오는 버그)
                    if any(p['row'] == r and p['col'] == c and p['direction'] == direction
                           for p in placed):
                        continue
                    # 3) 한 단어가 다른 단어를 완전히 포함(같은 방향)하는 경우 차단
                    #    ex) 'ES'(A,r,c)가 있을 때 'ESS'(A,r,c) → 위에서 이미 차단
                    #    역방향: 'ESS'가 먼저 있고 'ES'가 같은 시작 → 마지막 셀 뒤가 글자라 _can_place에서 차단됨
                    grid = self._place_word(grid, word, r, c, direction)
                    placed.append({'word': word, 'clue': clue,
                                   'row': r, 'col': c, 'direction': direction})
                    placed_ok = True
                    break

            if len(placed) > len(best_placed):
                best_placed = placed
                best_grid = copy.deepcopy(grid)
                if len(placed) >= min(len(wc_list), 12): break

        # 번호 부여
        if best_grid:
            self._assign_numbers(best_grid, best_placed)
        return best_grid, best_placed

    def _assign_numbers(self, grid, placed):
        numbered = {}
        num = 1
        for r in range(self.rows):
            for c in range(self.cols):
                if grid[r][c] == '#': continue
                starts = False
                for p in placed:
                    if p['row'] == r and p['col'] == c:
                        starts = True
                        break
                if starts:
                    numbered[(r, c)] = num
                    for p in placed:
                        if p['row'] == r and p['col'] == c:
                            p['number'] = num
                    num += 1
        self.numbered = numbered

# ════════════════════════════════════════════════════════════════
# 메인 앱
# ════════════════════════════════════════════════════════════════
class CrosswordApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("크로스워드 퍼즐 메이커 Pro - 한울교육훈련센터")
        self.geometry("1400x900")
        self.minsize(1100, 750)
        self.configure(fg_color=COLORS["bg"])

        # ── 아이콘 적용 (개발 환경 & PyInstaller --onefile exe 공용) ──
        _ico = resource_path("crossword.ico")
        if os.path.exists(_ico):
            self.iconbitmap(_ico)

        # State
        self.words_data = []          # [(word, clue, course, subject)]
        self.current_grid = None
        self.current_placed = []
        self.numbered = {}            # (r,c) -> number
        self.cell_widgets = {}        # (r,c) -> Entry widget
        self.cell_state = {}          # (r,c) -> 'empty'|'correct'|'wrong'
        self.grid_rows = 10
        self.grid_cols = 10
        self.difficulty_filter = "전체"
        self.selected_direction = 'A'
        self.selected_word_idx = -1
        self.selected_cell = None
        self.timer_running = False
        self.start_time = 0
        self.elapsed = 0
        self.score = 0
        self.hints_used = 0

        self._build_ui()

    # ── UI 구성 ────────────────────────────────────────────────
    def _build_ui(self):
        # 상단 타이틀 바
        self._build_titlebar()
        # 메인 컨테이너 (좌/중/우)
        main = ctk.CTkFrame(self, fg_color="transparent")
        main.pack(fill="both", expand=True, padx=10, pady=(0,10))
        main.columnconfigure(1, weight=1)
        main.rowconfigure(0, weight=1)

        self._build_left_panel(main)
        self._build_center_panel(main)
        self._build_right_panel(main)

    def _build_titlebar(self):
        bar = ctk.CTkFrame(self, fg_color=COLORS["panel"], height=60, corner_radius=0)
        bar.pack(fill="x", pady=(0,2))
        bar.pack_propagate(False)

        ctk.CTkLabel(bar, text="계통약어 학습용 크로스워드 퍼즐",
                     font=ctk.CTkFont(size=22, weight="bold"),
                     text_color=COLORS["gold"]).pack(side="left", padx=20)

        # 우측 버튼들
        btn_frame = ctk.CTkFrame(bar, fg_color="transparent")
        btn_frame.pack(side="right", padx=10)

        self._icon_btn(btn_frame, "📄 PDF 저장", self._export_pdf).pack(side="right", padx=4)
        self._icon_btn(btn_frame, "💾 저장", self._save_puzzle).pack(side="right", padx=4)
        self._icon_btn(btn_frame, "📂 불러오기", self._load_puzzle).pack(side="right", padx=4)

    def _icon_btn(self, parent, text, cmd, color=None):
        return ctk.CTkButton(parent, text=text, command=cmd,
                             width=110, height=32,
                             fg_color=color or COLORS["card"],
                             hover_color=COLORS["accent2"],
                             font=ctk.CTkFont(size=12))

    # ── 좌측 패널 ─────────────────────────────────────────────
    def _build_left_panel(self, parent):
        left = ctk.CTkFrame(parent, fg_color=COLORS["panel"], width=290, corner_radius=12)
        left.grid(row=0, column=0, sticky="nsew", padx=(0,6), pady=0)
        left.pack_propagate(False)

        # 엑셀 로드
        sec = self._section(left, "📁 데이터 로드")
        self._icon_btn(sec, "📊 엑셀 파일 열기", self._load_excel,
                       color=COLORS["accent"]).pack(fill="x", pady=4)
        self.file_label = ctk.CTkLabel(sec, text="파일 미선택",
                                        text_color=COLORS["subtext"],
                                        font=ctk.CTkFont(size=11), wraplength=240)
        self.file_label.pack(pady=2)

        self.word_count_label = ctk.CTkLabel(sec, text="단어: 0개",
                                              font=ctk.CTkFont(size=12, weight="bold"),
                                              text_color=COLORS["gold"])
        self.word_count_label.pack(pady=2)

        # 과정명 / 교과목 선택
        sec2 = self._section(left, "🎯 과정명 / 교과목 선택")

        # 과정명 드롭다운
        ctk.CTkLabel(sec2, text="과정명:", font=ctk.CTkFont(size=12),
                     text_color=COLORS["subtext"]).pack(anchor="w", padx=8, pady=(2,0))
        self.course_var = ctk.StringVar(value="전체")
        self.course_menu = ctk.CTkOptionMenu(
            sec2, variable=self.course_var,
            values=["전체"],
            command=self._on_course_change,
            fg_color=COLORS["card"],
            button_color=COLORS["accent2"],
            dropdown_fg_color=COLORS["panel"],
            width=240
        )
        self.course_menu.pack(fill="x", padx=8, pady=(2, 6))

        # 교과목 다중선택 (드롭다운 체크박스)
        ctk.CTkLabel(sec2, text="교과목 (다중선택):", font=ctk.CTkFont(size=12),
                     text_color=COLORS["subtext"]).pack(anchor="w", padx=8, pady=(2,0))

        # 교과목 선택 버튼 (클릭 시 팝업 패널 열림)
        self.subject_btn = ctk.CTkButton(
            sec2, text="▼  교과목 선택 (0개 선택)",
            command=self._open_subject_picker,
            fg_color=COLORS["card"],
            hover_color=COLORS["accent2"],
            font=ctk.CTkFont(size=12),
            anchor="w"
        )
        self.subject_btn.pack(fill="x", padx=8, pady=(2, 4))

        # 선택된 교과목 표시 레이블
        self.subject_info_label = ctk.CTkLabel(
            sec2, text="(과정명을 먼저 선택하세요)",
            text_color=COLORS["subtext"],
            font=ctk.CTkFont(size=10),
            wraplength=230
        )
        self.subject_info_label.pack(anchor="w", padx=8, pady=(0,4))

        # 내부 상태
        self.selected_subjects = set()   # 선택된 교과목명
        self.diff_var = ctk.StringVar(value="전체")  # 하위 호환용

        # 그리드 설정
        sec3 = self._section(left, "⚙️ 그리드 설정")
        row_frame = ctk.CTkFrame(sec3, fg_color="transparent")
        row_frame.pack(fill="x", pady=2)
        ctk.CTkLabel(row_frame, text="행:", width=40).pack(side="left")
        self.row_var = ctk.IntVar(value=10)
        ctk.CTkSlider(row_frame, from_=8, to=15, variable=self.row_var,
                      number_of_steps=7, width=140,
                      command=lambda v: self.row_val.configure(text=f"{int(v)}")).pack(side="left", padx=4)
        self.row_val = ctk.CTkLabel(row_frame, text="10", width=30)
        self.row_val.pack(side="left")

        col_frame = ctk.CTkFrame(sec3, fg_color="transparent")
        col_frame.pack(fill="x", pady=2)
        ctk.CTkLabel(col_frame, text="열:", width=40).pack(side="left")
        self.col_var = ctk.IntVar(value=10)
        ctk.CTkSlider(col_frame, from_=8, to=15, variable=self.col_var,
                      number_of_steps=7, width=140,
                      command=lambda v: self.col_val.configure(text=f"{int(v)}")).pack(side="left", padx=4)
        self.col_val = ctk.CTkLabel(col_frame, text="10", width=30)
        self.col_val.pack(side="left")

        # 단어 길이 필터
        len_frame = ctk.CTkFrame(sec3, fg_color="transparent")
        len_frame.pack(fill="x", pady=4)
        ctk.CTkLabel(len_frame, text="단어 길이:").pack(side="left")
        self.min_len = ctk.IntVar(value=2)
        self.max_len = ctk.IntVar(value=7)
        ctk.CTkLabel(len_frame, text="  최소").pack(side="left")
        ctk.CTkOptionMenu(len_frame, variable=self.min_len,
                          values=[str(i) for i in range(2,9)], width=55).pack(side="left", padx=2)
        ctk.CTkLabel(len_frame, text="최대").pack(side="left")
        ctk.CTkOptionMenu(len_frame, variable=self.max_len,
                          values=[str(i) for i in range(2,16)], width=55).pack(side="left", padx=2)

        # 생성 버튼
        ctk.CTkButton(left, text="🎲  퍼즐 생성!", command=self._generate_puzzle,
                      height=44, fg_color=COLORS["accent"],
                      hover_color="#c0392b",
                      font=ctk.CTkFont(size=16, weight="bold"),
                      corner_radius=10).pack(fill="x", padx=12, pady=8)

        # 단어 목록 (접기/펼치기 + 정답 숨김)
        sec4_header = ctk.CTkFrame(left, fg_color=COLORS["card"], corner_radius=8)
        sec4_header.pack(fill="x", padx=10, pady=5)
        sec4_top = ctk.CTkFrame(sec4_header, fg_color="transparent")
        sec4_top.pack(fill="x")

        ctk.CTkLabel(sec4_top, text="📋 단어 목록",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color=COLORS["cell_num"]).pack(side="left", padx=8, pady=(6,2))

        self.word_list_visible = True
        self.show_answers_var = ctk.BooleanVar(value=True)

        toggle_frame = ctk.CTkFrame(sec4_top, fg_color="transparent")
        toggle_frame.pack(side="right", padx=6, pady=2)

        ctk.CTkCheckBox(toggle_frame, text="정답숨김", variable=self.show_answers_var,
                        onvalue=True, offvalue=False,
                        command=self._update_word_list,
                        font=ctk.CTkFont(size=10),
                        checkbox_width=16, checkbox_height=16,
                        fg_color=COLORS["accent"]).pack(side="left", padx=4)

        self.toggle_btn = ctk.CTkButton(toggle_frame, text="▲ 접기",
                                         command=self._toggle_word_list,
                                         width=60, height=22,
                                         fg_color=COLORS["accent2"],
                                         font=ctk.CTkFont(size=11))
        self.toggle_btn.pack(side="left", padx=2)

        self.word_list_container = ctk.CTkFrame(sec4_header, fg_color="transparent")
        self.word_list_container.pack(fill="both", expand=True)

        self.word_list_box = ctk.CTkTextbox(self.word_list_container, height=180,
                                             fg_color=COLORS["empty"],
                                             text_color=COLORS["cell_text"],
                                             font=ctk.CTkFont(size=11))
        self.word_list_box.pack(fill="both", expand=True, padx=6, pady=(0,6))

    def _toggle_word_list(self):
        if self.word_list_visible:
            self.word_list_container.pack_forget()
            self.toggle_btn.configure(text="▼ 펼치기")
            self.word_list_visible = False
        else:
            self.word_list_container.pack(fill="both", expand=True)
            self.toggle_btn.configure(text="▲ 접기")
            self.word_list_visible = True

    # ── 중앙 패널 (퍼즐 그리드) ───────────────────────────────
    def _build_center_panel(self, parent):
        center = ctk.CTkFrame(parent, fg_color=COLORS["bg"], corner_radius=0)
        center.grid(row=0, column=1, sticky="nsew", padx=3)
        center.rowconfigure(1, weight=1)
        center.columnconfigure(0, weight=1)

        # 상태 바
        status = ctk.CTkFrame(center, fg_color=COLORS["panel"], height=45, corner_radius=8)
        status.grid(row=0, column=0, sticky="ew", pady=(0,6))
        status.pack_propagate(False)
        status.columnconfigure((0,1,2,3,4), weight=1)

        self.timer_label = ctk.CTkLabel(status, text="⏱️  00:00",
                                         font=ctk.CTkFont(size=16, weight="bold"),
                                         text_color=COLORS["gold"])
        self.timer_label.grid(row=0, column=0, padx=10)

        self.score_label = ctk.CTkLabel(status, text="🏆  0점",
                                         font=ctk.CTkFont(size=16, weight="bold"),
                                         text_color=COLORS["correct"])
        self.score_label.grid(row=0, column=1, padx=10)

        self.progress_label = ctk.CTkLabel(status, text="📊  0/0",
                                            font=ctk.CTkFont(size=14),
                                            text_color=COLORS["text"])
        self.progress_label.grid(row=0, column=2, padx=10)

        self.hint_label = ctk.CTkLabel(status, text="💡  힌트: 0",
                                        font=ctk.CTkFont(size=14),
                                        text_color=COLORS["subtext"])
        self.hint_label.grid(row=0, column=3, padx=10)

        btn_bar = ctk.CTkFrame(status, fg_color="transparent")
        btn_bar.grid(row=0, column=4, padx=8)
        ctk.CTkButton(btn_bar, text="✅ 채점", command=self._check_answers,
                       width=72, height=28, fg_color=COLORS["correct"]).pack(side="left", padx=2)
        ctk.CTkButton(btn_bar, text="💡 힌트", command=self._give_hint,
                       width=72, height=28, fg_color=COLORS["gold"],
                       text_color="#000").pack(side="left", padx=2)
        ctk.CTkButton(btn_bar, text="🔄 초기화", command=self._reset_puzzle,
                       width=72, height=28, fg_color=COLORS["card"]).pack(side="left", padx=2)
        ctk.CTkButton(btn_bar, text="👁️ 정답", command=self._reveal_all,
                       width=72, height=28, fg_color=COLORS["accent2"]).pack(side="left", padx=2)

        # 퍼즐 캔버스 영역
        canvas_outer = ctk.CTkFrame(center, fg_color=COLORS["panel"], corner_radius=10)
        canvas_outer.grid(row=1, column=0, sticky="nsew")
        canvas_outer.rowconfigure(0, weight=1)
        canvas_outer.columnconfigure(0, weight=1)

        self.canvas_frame = ctk.CTkScrollableFrame(canvas_outer,
                                                    fg_color=COLORS["empty"],
                                                    corner_radius=8)
        self.canvas_frame.grid(row=0, column=0, sticky="nsew", padx=4, pady=4)

        self.puzzle_placeholder = ctk.CTkLabel(
            self.canvas_frame,
            text="퍼즐을 생성하려면\n왼쪽에서 엑셀 파일을 열고\n'퍼즐 생성!' 버튼을 클릭하세요.",
            font=ctk.CTkFont(size=18),
            text_color=COLORS["subtext"]
        )
        self.puzzle_placeholder.pack(expand=True, pady=80)

    # ── 우측 패널 (단서 목록) ────────────────────────────────
    def _build_right_panel(self, parent):
        right = ctk.CTkFrame(parent, fg_color=COLORS["panel"], width=280, corner_radius=12)
        right.grid(row=0, column=2, sticky="nsew", padx=(6,0))
        right.pack_propagate(False)

        ctk.CTkLabel(right, text="📖 단서",
                     font=ctk.CTkFont(size=16, weight="bold"),
                     text_color=COLORS["gold"]).pack(pady=(12,4))

        # Across / Down 탭
        tab = ctk.CTkTabview(right, fg_color=COLORS["card"],
                              segmented_button_fg_color=COLORS["empty"],
                              segmented_button_selected_color=COLORS["accent"])
        tab.pack(fill="both", expand=True, padx=8, pady=4)
        self.across_tab = tab.add("→ 가로")
        self.down_tab   = tab.add("↓ 세로")

        self.across_clue_frame = ctk.CTkScrollableFrame(self.across_tab,
                                                         fg_color="transparent")
        self.across_clue_frame.pack(fill="both", expand=True)

        self.down_clue_frame = ctk.CTkScrollableFrame(self.down_tab,
                                                       fg_color="transparent")
        self.down_clue_frame.pack(fill="both", expand=True)

        # 정답 현황
        sec = self._section(right, "📊 현황")
        self.progress_bar = ctk.CTkProgressBar(sec, fg_color=COLORS["empty"],
                                                progress_color=COLORS["correct"])
        self.progress_bar.set(0)
        self.progress_bar.pack(fill="x", pady=4)
        self.progress_detail = ctk.CTkLabel(sec, text="아직 시작 전",
                                             text_color=COLORS["subtext"],
                                             font=ctk.CTkFont(size=12))
        self.progress_detail.pack()

        # 통계
        sec2 = self._section(right, "📈 통계")
        self.stats_label = ctk.CTkLabel(sec2, text="퍼즐을 생성해 주세요",
                                         text_color=COLORS["subtext"],
                                         font=ctk.CTkFont(size=11),
                                         justify="left")
        self.stats_label.pack(anchor="w", padx=4)

    # ── 헬퍼 ─────────────────────────────────────────────────
    def _section(self, parent, title):
        frame = ctk.CTkFrame(parent, fg_color=COLORS["card"], corner_radius=8)
        frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(frame, text=title,
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color=COLORS["cell_num"]).pack(anchor="w", padx=8, pady=(6,2))
        return frame

    # ════════════════════════════════════════════════════════
    # 엑셀 로드
    # ════════════════════════════════════════════════════════
    def _load_excel(self):
        path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not path: return
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            self.words_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[1]: continue
                clue = str(row[0]).strip()
                word = str(row[1]).strip().upper()
                diff = str(row[2]).strip() if len(row) > 2 and row[2] else "중"
                if word.isalpha():
                    self.words_data.append((word, clue, diff))

            fname = os.path.basename(path)
            self.file_label.configure(text=fname)
            self.word_count_label.configure(text=f"단어: {len(self.words_data)}개")
            self._update_word_list()
            messagebox.showinfo("로드 완료", f"총 {len(self.words_data)}개 단어를 불러왔습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"파일 읽기 실패:\n{e}")

    # ════════════════════════════════════════════════════════
    # 엑셀 로드
    # ════════════════════════════════════════════════════════
    def _load_excel(self):
        path = filedialog.askopenfilename(
            title="엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if not path: return
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            self.words_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0] or not row[1]: continue
                clue    = str(row[0]).strip()
                word    = str(row[1]).strip().upper()
                course  = str(row[2]).strip() if len(row) > 2 and row[2] else "미분류"
                subject = str(row[3]).strip() if len(row) > 3 and row[3] else "미분류"
                if word.isalpha():
                    self.words_data.append((word, clue, course, subject))

            fname = os.path.basename(path)
            self.file_label.configure(text=fname)
            self.word_count_label.configure(text=f"단어: {len(self.words_data)}개")
            self._refresh_course_menu()
            self._update_word_list()
            messagebox.showinfo("로드 완료", f"총 {len(self.words_data)}개 단어를 불러왔습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"파일 읽기 실패:\n{e}")

    def _refresh_course_menu(self):
        """과정명 드롭다운 목록 갱신"""
        courses = sorted(set(d[2] for d in self.words_data))
        values = ["전체"] + courses
        self.course_menu.configure(values=values)
        self.course_var.set("전체")
        self.selected_subjects.clear()
        self._update_subject_btn_label()

    def _on_course_change(self, value=None):
        """과정명 변경 시 교과목 선택 초기화"""
        self.selected_subjects.clear()
        self._update_subject_btn_label()
        self._update_word_list()

    def _get_subjects_for_course(self):
        """현재 선택된 과정명에 해당하는 교과목 목록 반환"""
        course = self.course_var.get()
        if course == "전체":
            subjects = sorted(set(d[3] for d in self.words_data))
        else:
            subjects = sorted(set(d[3] for d in self.words_data if d[2] == course))
        return subjects

    def _open_subject_picker(self):
        """교과목 체크박스 팝업 창"""
        subjects = self._get_subjects_for_course()
        if not subjects:
            messagebox.showinfo("알림", "먼저 엑셀 파일을 로드하세요.")
            return

        win = tk.Toplevel(self)
        win.title("교과목 선택")
        win.configure(bg=COLORS["bg"])
        win.resizable(False, False)

        # 창 위치: 부모 창 기준
        x = self.winfo_x() + 100
        y = self.winfo_y() + 200
        win.geometry(f"320x500+{x}+{y}")
        win.grab_set()

        # 헤더
        hdr = tk.Frame(win, bg=COLORS["panel"], height=40)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="📚 교과목 선택", bg=COLORS["panel"],
                 fg=COLORS["gold"], font=("Arial", 13, "bold")).pack(side="left", padx=12, pady=8)

        # 전체선택 / 전체해제 버튼
        btn_bar = tk.Frame(win, bg=COLORS["card"])
        btn_bar.pack(fill="x", padx=8, pady=4)
        check_vars = {}

        def select_all():
            for var in check_vars.values():
                var.set(True)

        def deselect_all():
            for var in check_vars.values():
                var.set(False)

        tk.Button(btn_bar, text="전체 선택", bg=COLORS["accent2"], fg="white",
                  bd=0, padx=8, command=select_all).pack(side="left", padx=4, pady=4)
        tk.Button(btn_bar, text="전체 해제", bg=COLORS["card"], fg=COLORS["text"],
                  bd=0, padx=8, command=deselect_all).pack(side="left", padx=4, pady=4)

        # 검색 필터
        search_var = tk.StringVar()
        search_entry = tk.Entry(btn_bar, textvariable=search_var,
                                bg=COLORS["empty"], fg=COLORS["text"],
                                insertbackground=COLORS["accent"],
                                bd=1, relief="flat", width=14)
        search_entry.pack(side="right", padx=6, pady=4)
        tk.Label(btn_bar, text="🔍", bg=COLORS["card"],
                 fg=COLORS["subtext"]).pack(side="right")

        # 스크롤 영역
        scroll_frame_outer = tk.Frame(win, bg=COLORS["bg"])
        scroll_frame_outer.pack(fill="both", expand=True, padx=8, pady=4)

        canvas = tk.Canvas(scroll_frame_outer, bg=COLORS["bg"], bd=0, highlightthickness=0)
        scrollbar = tk.Scrollbar(scroll_frame_outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        inner = tk.Frame(canvas, bg=COLORS["bg"])
        canvas_window = canvas.create_window((0, 0), window=inner, anchor="nw")

        def on_resize(event):
            canvas.itemconfig(canvas_window, width=event.width)
        canvas.bind("<Configure>", on_resize)

        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        inner.bind("<Configure>", on_frame_configure)

        # 마우스 휠 스크롤
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # 체크박스 생성
        all_check_frames = []

        def build_checkboxes(filter_text=""):
            for w in inner.winfo_children():
                w.destroy()
            all_check_frames.clear()
            for subj in subjects:
                if filter_text and filter_text.lower() not in subj.lower():
                    continue
                var = check_vars.setdefault(subj, tk.BooleanVar(value=(subj in self.selected_subjects)))
                f = tk.Frame(inner, bg=COLORS["bg"])
                f.pack(fill="x", padx=4, pady=1)
                cb = tk.Checkbutton(f, text=subj, variable=var,
                                    bg=COLORS["bg"], fg=COLORS["text"],
                                    selectcolor=COLORS["card"],
                                    activebackground=COLORS["bg"],
                                    activeforeground=COLORS["gold"],
                                    anchor="w", font=("Arial", 11))
                cb.pack(fill="x")
                all_check_frames.append(f)

        build_checkboxes()

        def on_search_change(*args):
            build_checkboxes(search_var.get())
        search_var.trace_add("write", on_search_change)

        # 확인 버튼
        def confirm():
            self.selected_subjects = {subj for subj, var in check_vars.items() if var.get()}
            self._update_subject_btn_label()
            self._update_word_list()
            canvas.unbind_all("<MouseWheel>")
            win.destroy()

        def on_close():
            canvas.unbind_all("<MouseWheel>")
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", on_close)

        foot = tk.Frame(win, bg=COLORS["panel"], height=44)
        foot.pack(fill="x", side="bottom")
        foot.pack_propagate(False)
        tk.Button(foot, text="✅  확인", bg=COLORS["accent"], fg="white",
                  font=("Arial", 12, "bold"), bd=0, padx=20,
                  command=confirm).pack(pady=8)

    def _update_subject_btn_label(self):
        cnt = len(self.selected_subjects)
        if cnt == 0:
            self.subject_btn.configure(text="▼  교과목 선택 (전체)")
            self.subject_info_label.configure(text="(전체 교과목 포함)")
        else:
            preview = ", ".join(sorted(self.selected_subjects)[:3])
            if cnt > 3:
                preview += f" 외 {cnt-3}개"
            self.subject_btn.configure(text=f"▼  교과목 선택 ({cnt}개 선택)")
            self.subject_info_label.configure(text=preview)

    def _update_word_list(self):
        self.word_list_box.configure(state="normal")
        self.word_list_box.delete("1.0", "end")

        # 교과목 미선택 시 안내 문구
        if not self.selected_subjects:
            course = self.course_var.get()
            if course == "전체" or not self.words_data:
                self.word_list_box.insert("end", "📌 과정명을 선택 후\n교과목을 선택하세요.\n")
            else:
                self.word_list_box.insert("end", f"📌 '{course}' 과정의\n교과목을 선택하세요.\n")
            self.word_list_box.configure(state="disabled")
            return

        filtered = self._get_filtered_words()
        hide = self.show_answers_var.get()
        if filtered:
            for w, c, course, subj in filtered:
                answer = "?" * len(w) if hide else w
                self.word_list_box.insert("end", f"[{subj}] {answer} - {c}\n")
        else:
            self.word_list_box.insert("end", "선택한 조건에 해당하는\n단어가 없습니다.\n")
        self.word_list_box.configure(state="disabled")

    def _get_filtered_words(self):
        course = self.course_var.get()
        min_l = self.min_len.get()
        max_l = self.max_len.get()
        result = []
        for item in self.words_data:
            w, c, crs, subj = item
            # 과정명 필터
            if course != "전체" and crs != course:
                continue
            # 교과목 필터 (선택된 교과목이 없으면 빈 결과 반환 — 단어 목록/생성 모두 막힘)
            if subj not in self.selected_subjects:
                continue
            # 단어 길이 필터
            if not (min_l <= len(w) <= max_l):
                continue
            result.append((w, c, crs, subj))
        return result

    def _on_diff_change(self):
        self._update_word_list()

    # ════════════════════════════════════════════════════════
    # 퍼즐 생성
    # ════════════════════════════════════════════════════════
    def _generate_puzzle(self):
        filtered = self._get_filtered_words()
        if len(filtered) < 3:
            messagebox.showwarning("단어 부족",
                "최소 3개 이상의 단어가 필요합니다.\n엑셀 파일을 로드하거나 필터를 확인하세요.")
            return

        self.grid_rows = self.row_var.get()
        self.grid_cols = self.col_var.get()

        wc_list = [(w, c) for w, c, crs, subj in filtered]
        gen = CrosswordGenerator(wc_list, self.grid_rows, self.grid_cols, max_attempts=60)

        # 생성 중 메시지 (placeholder가 살아있을 때만)
        if self.puzzle_placeholder.winfo_exists():
            self.puzzle_placeholder.configure(text="⏳ 퍼즐 생성 중...")
        self.update()

        grid, placed = gen.generate()
        self.numbered = gen.numbered if hasattr(gen, 'numbered') else {}

        if not grid or not placed:
            messagebox.showwarning("생성 실패", "단어 배치에 실패했습니다. 그리드 크기를 늘리거나 다시 시도하세요.")
            if self.puzzle_placeholder.winfo_exists():
                self.puzzle_placeholder.configure(text="생성 실패 - 다시 시도해주세요")
            return

        self.current_grid = grid
        self.current_placed = placed
        self.cell_state = {}
        self.hints_used = 0
        self.score = 0
        self.selected_word_idx = -1
        self.selected_direction = 'A'
        self.selected_cell = None

        self._draw_puzzle()
        self._build_clue_list()
        self._update_stats()
        self._start_timer()

    # ════════════════════════════════════════════════════════
    # 그리드 그리기
    # ════════════════════════════════════════════════════════
    def _draw_puzzle(self):
        # 기존 위젯 제거
        for w in self.canvas_frame.winfo_children():
            w.destroy()
        self.cell_widgets = {}

        if not self.current_grid: return

        CELL_SIZE = max(34, min(52, 640 // max(self.grid_rows, self.grid_cols)))
        NUM_FONT_SIZE = max(7, CELL_SIZE // 4)
        LETTER_FONT_SIZE = max(11, int(CELL_SIZE * 0.45))

        outer = tk.Frame(self.canvas_frame, bg=COLORS["empty"])
        outer.pack(expand=True, padx=10, pady=10)

        for r in range(self.grid_rows):
            for c in range(self.grid_cols):
                cell_val = self.current_grid[r][c]
                if cell_val == '#':
                    lbl = tk.Frame(outer, width=CELL_SIZE, height=CELL_SIZE,
                                   bg=COLORS["empty"], bd=0)
                    lbl.grid(row=r, column=c, padx=1, pady=1)
                    lbl.grid_propagate(False)
                else:
                    # 셀 컨테이너 (Canvas 기반 → 번호+입력 겹침 없이 깔끔하게)
                    cell = tk.Canvas(outer, width=CELL_SIZE, height=CELL_SIZE,
                                     bg=COLORS["cell_bg"],
                                     bd=0, highlightthickness=1,
                                     highlightbackground=COLORS["border"])
                    cell.grid(row=r, column=c, padx=1, pady=1)

                    # ── 셀 번호 (좌상단, Canvas text로 그려서 항상 보임) ──
                    if (r, c) in self.numbered:
                        cell.create_text(
                            2, 1,
                            text=str(self.numbered[(r, c)]),
                            anchor="nw",
                            font=("Arial", NUM_FONT_SIZE, "bold"),
                            fill="#a0c4ff",
                            tags="num"
                        )

                    # ── 입력 Entry (번호 아래쪽 공간 차지) ──
                    num_offset = NUM_FONT_SIZE + 2 if (r, c) in self.numbered else 0
                    var = tk.StringVar()
                    entry = tk.Entry(cell, textvariable=var,
                                     font=("Arial", LETTER_FONT_SIZE, "bold"),
                                     bg=COLORS["cell_bg"],
                                     fg=COLORS["cell_text"],
                                     insertbackground=COLORS["accent"],
                                     bd=0, highlightthickness=0,
                                     justify="center",
                                     width=1)
                    # Entry를 번호 아래에 배치 (번호 없는 셀은 중앙)
                    ey = (CELL_SIZE + num_offset) // 2
                    cell.create_window(CELL_SIZE // 2, ey,
                                       window=entry,
                                       width=CELL_SIZE - 4,
                                       height=CELL_SIZE - num_offset - 4)

                    entry.bind("<KeyRelease>", lambda e, row=r, col=c, v=var: self._on_key(e, row, col, v))
                    entry.bind("<FocusIn>",    lambda e, row=r, col=c: self._on_focus(row, col))
                    var.trace_add("write", lambda *a, row=r, col=c, v=var: self._on_entry_change(row, col, v))

                    self.cell_widgets[(r, c)] = (entry, var, cell)

    def _on_key(self, event, row, col, var):
        val = var.get().upper()
        if len(val) > 1:
            var.set(val[-1])
        elif val and val.isalpha():
            var.set(val)
            self._move_next(row, col)
        elif event.keysym == 'BackSpace':
            var.set('')
        elif event.keysym in ('Left', 'Right', 'Up', 'Down'):
            self._navigate(row, col, event.keysym)

    def _move_next(self, row, col):
        # 현재 선택 방향으로 이동
        if self.selected_direction == 'A':
            next_cell = (row, col+1)
        else:
            next_cell = (row+1, col)
        if next_cell in self.cell_widgets:
            self.cell_widgets[next_cell][0].focus_set()

    def _navigate(self, row, col, key):
        moves = {'Left': (0,-1), 'Right': (0,1), 'Up': (-1,0), 'Down': (1,0)}
        dr, dc = moves.get(key, (0,0))
        target = (row+dr, col+dc)
        if target in self.cell_widgets:
            self.cell_widgets[target][0].focus_set()

    def _on_focus(self, row, col):
        toggle = (self.selected_cell == (row, col))
        self.selected_cell = (row, col)
        self._highlight_word(row, col, toggle=toggle)

    def _on_entry_change(self, row, col, var):
        self._update_progress()

    def _set_cell_bg(self, r, c, color):
        """Canvas 셀과 그 안의 Entry 배경색을 함께 변경"""
        if (r, c) not in self.cell_widgets: return
        entry, var, canvas = self.cell_widgets[(r, c)]
        canvas.configure(bg=color)
        try: canvas.itemconfigure("num", fill="#a0c4ff")
        except: pass
        entry.configure(bg=color)

    def _highlight_word(self, row, col, toggle=False):
        # 이 셀이 속한 가로/세로 단어 모두 찾기
        matches = {}  # direction -> (p, cells)
        for p in self.current_placed:
            wr, wc, wd = p['row'], p['col'], p['direction']
            cells = [(wr, wc+i) if wd == 'A' else (wr+i, wc) for i in range(len(p['word']))]
            if (row, col) in cells:
                matches[wd] = (p, cells)

        if not matches:
            return

        # 방향 결정: toggle 시 반대 방향으로, 없으면 현재 방향 유지
        if toggle and len(matches) == 2:
            new_dir = 'D' if self.selected_direction == 'A' else 'A'
        elif self.selected_direction in matches:
            new_dir = self.selected_direction
        else:
            new_dir = next(iter(matches))

        p, cells = matches[new_dir]

        # 모든 셀 초기화
        for (r, c) in self.cell_widgets:
            self._set_cell_bg(r, c, COLORS["cell_bg"])

        self.selected_direction = new_dir
        self.selected_word_idx = self.current_placed.index(p)
        for cr, cc in cells:
            self._set_cell_bg(cr, cc, COLORS["highlight"])
        self._set_cell_bg(row, col, "#e94580")

    # ════════════════════════════════════════════════════════
    # 단서 목록
    # ════════════════════════════════════════════════════════
    def _build_clue_list(self):
        for w in self.across_clue_frame.winfo_children(): w.destroy()
        for w in self.down_clue_frame.winfo_children(): w.destroy()

        across = sorted([p for p in self.current_placed if p['direction'] == 'A'], key=lambda x: x.get('number', 99))
        down   = sorted([p for p in self.current_placed if p['direction'] == 'D'], key=lambda x: x.get('number', 99))

        for p in across:
            self._clue_item(self.across_clue_frame, p)
        for p in down:
            self._clue_item(self.down_clue_frame, p)

    def _clue_item(self, parent, p):
        num = p.get('number', '?')
        word = p['word']
        clue = p['clue']
        direction = p['direction']
        length = len(word)

        frame = ctk.CTkFrame(parent, fg_color=COLORS["card"], corner_radius=6)
        frame.pack(fill="x", pady=2, padx=2)
        frame.bind("<Button-1>", lambda e, pp=p: self._jump_to_word(pp))

        color = COLORS["across"] if direction == 'A' else COLORS["down"]
        ctk.CTkLabel(frame,
                     text=f"{'→' if direction=='A' else '↓'} {num}",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color=COLORS["gold"],
                     width=36).pack(side="left", padx=(6,2), pady=4)

        ctk.CTkLabel(frame,
                     text=f"{clue}  ({length}글자)",
                     font=ctk.CTkFont(size=12),
                     text_color=COLORS["text"],
                     anchor="w",
                     wraplength=180).pack(side="left", fill="x", expand=True, padx=4, pady=4)

    def _jump_to_word(self, p):
        wr, wc = p['row'], p['col']
        if (wr, wc) in self.cell_widgets:
            self.cell_widgets[(wr, wc)][0].focus_set()

    # ════════════════════════════════════════════════════════
    # 채점 / 힌트 / 초기화
    # ════════════════════════════════════════════════════════
    def _check_answers(self):
        if not self.current_placed: return
        correct_words = 0
        total_words = len(self.current_placed)

        for p in self.current_placed:
            word = p['word']
            wr, wc, wd = p['row'], p['col'], p['direction']
            user_word = ""
            cells = []
            for i in range(len(word)):
                cr, cc = (wr, wc+i) if wd == 'A' else (wr+i, wc)
                cells.append((cr, cc))
                if (cr, cc) in self.cell_widgets:
                    v = self.cell_widgets[(cr, cc)][1].get().upper()
                    user_word += v
                else:
                    user_word += '?'

            is_correct = user_word == word
            if is_correct:
                correct_words += 1

            for i, (cr, cc) in enumerate(cells):
                bg = COLORS["correct"] if is_correct else COLORS["wrong"]
                self._set_cell_bg(cr, cc, bg)

        self._stop_timer()
        base_score = correct_words * 10
        time_bonus = max(0, 300 - self.elapsed) // 10
        hint_penalty = self.hints_used * 5
        self.score = max(0, base_score + time_bonus - hint_penalty)
        self.score_label.configure(text=f"🏆  {self.score}점")

        # 완료 팝업
        elapsed_str = self._fmt_time(self.elapsed)
        msg = (f"✅ 채점 완료!\n\n"
               f"정답 단어: {correct_words}/{total_words}\n"
               f"소요 시간: {elapsed_str}\n"
               f"힌트 사용: {self.hints_used}회\n\n"
               f"🏆 최종 점수: {self.score}점")
        messagebox.showinfo("채점 결과", msg)
        self._update_progress()

    def _give_hint(self):
        if not self.current_placed: return

        idx = self.selected_word_idx
        if not (0 <= idx < len(self.current_placed)):
            messagebox.showinfo("힌트", "먼저 셀을 클릭해 단어를 선택하세요.")
            return

        p = self.current_placed[idx]
        word = p['word']
        wr, wc, wd = p['row'], p['col'], p['direction']

        # 빈 칸 또는 틀린 글자 찾아서 수정
        for i, ch in enumerate(word):
            cr, cc = (wr, wc+i) if wd == 'A' else (wr+i, wc)
            if (cr, cc) in self.cell_widgets:
                v = self.cell_widgets[(cr, cc)][1].get().upper()
                if v != ch:
                    self.cell_widgets[(cr, cc)][1].set(ch)
                    self._set_cell_bg(cr, cc, COLORS["gold"])
                    self.hints_used += 1
                    self.hint_label.configure(text=f"💡  힌트: {self.hints_used}")
                    return

        # 해당 단어가 이미 완성된 경우
        d_str = "가로" if wd == 'A' else "세로"
        messagebox.showinfo("힌트", f"{p.get('number','?')}번 {d_str} 단어는 이미 완성됐습니다!")

    def _reset_puzzle(self):
        for (r, c) in list(self.cell_widgets.keys()):
            self.cell_widgets[(r, c)][1].set('')
            self._set_cell_bg(r, c, COLORS["cell_bg"])
        self.hints_used = 0
        self.hint_label.configure(text="💡  힌트: 0")
        self.score = 0
        self.score_label.configure(text="🏆  0점")
        self._start_timer()
        self._update_progress()

    def _reveal_all(self):
        if not messagebox.askyesno("정답 공개", "모든 정답을 보시겠습니까?\n(점수가 0점이 됩니다)"):
            return
        for p in self.current_placed:
            word = p['word']
            wr, wc, wd = p['row'], p['col'], p['direction']
            for i, ch in enumerate(word):
                cr, cc = (wr, wc+i) if wd == 'A' else (wr+i, wc)
                if (cr, cc) in self.cell_widgets:
                    self.cell_widgets[(cr, cc)][1].set(ch)
                    self._set_cell_bg(cr, cc, COLORS["accent2"])
        self.score = 0
        self.score_label.configure(text="🏆  0점")
        self._stop_timer()

    # ════════════════════════════════════════════════════════
    # 타이머
    # ════════════════════════════════════════════════════════
    def _start_timer(self):
        self.timer_running = True
        self.start_time = time.time()
        self.elapsed = 0
        self._tick()

    def _stop_timer(self):
        self.timer_running = False

    def _tick(self):
        if not self.timer_running: return
        self.elapsed = int(time.time() - self.start_time)
        self.timer_label.configure(text=f"⏱️  {self._fmt_time(self.elapsed)}")
        self.after(1000, self._tick)

    def _fmt_time(self, secs):
        return f"{secs//60:02d}:{secs%60:02d}"

    # ════════════════════════════════════════════════════════
    # 진행 상황 업데이트
    # ════════════════════════════════════════════════════════
    def _update_progress(self):
        if not self.current_placed: return
        total_cells = sum(len(p['word']) for p in self.current_placed)
        filled = 0
        for (r, c), (entry, var, cell) in self.cell_widgets.items():
            if var.get().strip(): filled += 1

        pct = filled / total_cells if total_cells else 0
        self.progress_bar.set(pct)
        self.progress_label.configure(text=f"📊  {filled}/{total_cells} 칸")
        self.progress_detail.configure(text=f"완성률 {pct*100:.0f}%")

    def _update_stats(self):
        total = len(self.current_placed)
        across = sum(1 for p in self.current_placed if p['direction'] == 'A')
        down   = total - across
        total_cells = sum(len(p['word']) for p in self.current_placed)
        words = [p['word'] for p in self.current_placed]
        avg_len = sum(len(w) for w in words) / len(words) if words else 0

        self.stats_label.configure(
            text=(f"가로 단어: {across}개\n"
                  f"세로 단어: {down}개\n"
                  f"총 단어: {total}개\n"
                  f"총 칸 수: {total_cells}개\n"
                  f"평균 단어 길이: {avg_len:.1f}자")
        )
        self.progress_label.configure(text=f"📊  0/{total_cells}")

    # ════════════════════════════════════════════════════════
    # 저장 / 불러오기
    # ════════════════════════════════════════════════════════
    def _save_puzzle(self):
        if not self.current_grid:
            messagebox.showwarning("저장 실패", "생성된 퍼즐이 없습니다.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="퍼즐 저장"
        )
        if not path: return
        data = {
            "grid": self.current_grid,
            "placed": self.current_placed,
            "numbered": {f"{k[0]},{k[1]}": v for k, v in self.numbered.items()},
            "rows": self.grid_rows,
            "cols": self.grid_cols,
            "words_data": [
                {"word": w, "clue": c, "course": crs, "subject": subj}
                for w, c, crs, subj in self.words_data
            ],
        }
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        messagebox.showinfo("저장 완료", f"퍼즐이 저장되었습니다:\n{path}")

    def _load_puzzle(self):
        path = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")],
            title="퍼즐 불러오기"
        )
        if not path: return
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            self.current_grid = data["grid"]
            self.current_placed = data["placed"]
            self.numbered = {tuple(int(x) for x in k.split(',')): v
                             for k, v in data.get("numbered", {}).items()}

            # 그리드 크기 복원 + 슬라이더/레이블 동기화 (범위 8~15로 클램핑)
            self.grid_rows = max(8, min(15, data["rows"]))
            self.grid_cols = max(8, min(15, data["cols"]))
            self.row_var.set(self.grid_rows)
            self.col_var.set(self.grid_cols)
            self.row_val.configure(text=str(self.grid_rows))
            self.col_val.configure(text=str(self.grid_cols))

            # words_data 복원 (신규 포맷: course/subject, 구버전: diff)
            raw_words = data.get("words_data", [])
            if raw_words:
                self.words_data = []
                for item in raw_words:
                    w   = item["word"]
                    c   = item["clue"]
                    # 구버전(diff) → course/subject 호환
                    crs  = item.get("course", item.get("diff", "미분류"))
                    subj = item.get("subject", "미분류")
                    self.words_data.append((w, c, crs, subj))
                fname = os.path.basename(path)
                self.file_label.configure(text=f"[JSON] {fname}")
                self.word_count_label.configure(text=f"단어: {len(self.words_data)}개")
                self._refresh_course_menu()
                self._update_word_list()

            self.cell_state = {}
            self.hints_used = 0
            self._draw_puzzle()
            self._build_clue_list()
            self._update_stats()
            self._start_timer()
            messagebox.showinfo("불러오기 완료", "퍼즐을 불러왔습니다!")
        except Exception as e:
            messagebox.showerror("오류", f"불러오기 실패:\n{e}")

    # ════════════════════════════════════════════════════════
    # PDF 내보내기
    # ════════════════════════════════════════════════════════
    def _export_pdf(self):
        if not self.current_grid:
            messagebox.showwarning("저장 실패", "생성된 퍼즐이 없습니다.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            title="PDF로 저장"
        )
        if not path: return
        try:
            self._create_pdf(path)
            messagebox.showinfo("저장 완료", f"PDF가 저장되었습니다:\n{path}")
        except Exception as e:
            messagebox.showerror("오류", f"PDF 생성 실패:\n{e}")

    def _create_pdf(self, path):
        """3페이지 PDF 생성
        1페이지: 퍼즐 그리드
        2페이지: 단서 (가로 / 세로)
        3페이지: 정답
        """
        from reportlab.platypus import PageBreak, KeepInFrame

        PAGE_W, PAGE_H = A4
        doc = SimpleDocTemplate(path, pagesize=A4,
                                leftMargin=15*mm, rightMargin=15*mm,
                                topMargin=15*mm, bottomMargin=15*mm)

        # ── 공통 스타일 ─────────────────────────────────────
        title_style = ParagraphStyle(
            'kor_title',
            fontName=KOR_FONT_BOLD, fontSize=22,
            textColor=colors.HexColor('#0f3460'),
            spaceAfter=3, alignment=TA_CENTER, leading=28,
        )
        page_title_style = ParagraphStyle(
            'kor_page_title',
            fontName=KOR_FONT_BOLD, fontSize=16,
            textColor=colors.HexColor('#0f3460'),
            spaceAfter=3, alignment=TA_CENTER, leading=22,
        )
        sub_style = ParagraphStyle(
            'kor_sub',
            fontName=KOR_FONT, fontSize=10,
            textColor=colors.HexColor('#555555'),
            spaceAfter=8, alignment=TA_CENTER,
        )
        heading_style = ParagraphStyle(
            'kor_heading',
            fontName=KOR_FONT_BOLD, fontSize=13,
            textColor=colors.HexColor('#0f3460'),
            spaceAfter=4, spaceBefore=8,
        )
        clue_style = ParagraphStyle(
            'kor_clue',
            fontName=KOR_FONT, fontSize=10,
            textColor=colors.HexColor('#222222'),
            spaceAfter=3, leftIndent=8, leading=14,
        )
        answer_style = ParagraphStyle(
            'kor_answer',
            fontName=KOR_FONT_BOLD, fontSize=10,
            textColor=colors.HexColor('#0f3460'),
            spaceAfter=2, leftIndent=8,
        )
        num_cell_style = ParagraphStyle(
            'kor_numcell',
            fontName=KOR_FONT, fontSize=6,
            textColor=colors.HexColor('#003399'),
            leading=7,
        )

        # ── 공통 정보 ────────────────────────────────────────
        course_label = self.course_var.get()
        subj_cnt = len(self.selected_subjects)
        subj_label = f"{subj_cnt}개 교과목" if subj_cnt > 0 else "전체 교과목"
        info_text = (f"과정명: {course_label}  |  {subj_label}  |  "
                     f"단어 수: {len(self.current_placed)}개  |  "
                     f"그리드: {self.grid_rows}×{self.grid_cols}")

        rows, cols = self.grid_rows, self.grid_cols
        avail_w = PAGE_W - 30*mm
        CELL_MM = min(16, avail_w / cols / mm)

        across = sorted([p for p in self.current_placed if p['direction'] == 'A'],
                        key=lambda x: x.get('number', 99))
        down   = sorted([p for p in self.current_placed if p['direction'] == 'D'],
                        key=lambda x: x.get('number', 99))

        story = []

        # ════════════════════════════════════════════════════
        # 1페이지: 퍼즐 그리드
        # ════════════════════════════════════════════════════
        story.append(Paragraph("크로스워드 퍼즐", title_style))
        story.append(Paragraph(info_text, sub_style))
        story.append(Spacer(1, 6*mm))

        # 그리드 테이블
        table_data = []
        for r in range(rows):
            row_data = []
            for c in range(cols):
                val = self.current_grid[r][c]
                if val == '#':
                    row_data.append('')
                else:
                    num_str = str(self.numbered[(r, c)]) if (r, c) in self.numbered else ''
                    row_data.append(Paragraph(num_str, num_cell_style))
            table_data.append(row_data)

        col_widths  = [CELL_MM * mm] * cols
        row_heights = [CELL_MM * mm] * rows
        grid_table  = Table(table_data, colWidths=col_widths, rowHeights=row_heights)

        ts = TableStyle([
            ('GRID',            (0,0), (-1,-1), 0.7, colors.HexColor('#2d4a7a')),
            ('VALIGN',          (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING',     (0,0), (-1,-1), 1),
            ('TOPPADDING',      (0,0), (-1,-1), 1),
            ('RIGHTPADDING',    (0,0), (-1,-1), 0),
            ('BOTTOMPADDING',   (0,0), (-1,-1), 0),
        ])
        for r in range(rows):
            for c in range(cols):
                if self.current_grid[r][c] == '#':
                    ts.add('BACKGROUND', (c,r), (c,r), colors.HexColor('#1a1a2e'))
                    ts.add('GRID',       (c,r), (c,r), 0, colors.HexColor('#1a1a2e'))
                else:
                    ts.add('BACKGROUND', (c,r), (c,r), colors.white)
        grid_table.setStyle(ts)

        # 그리드를 페이지 중앙에 배치
        grid_w = CELL_MM * mm * cols
        grid_offset = (avail_w - grid_w) / 2
        if grid_offset > 0:
            centered = Table([[grid_table]], colWidths=[avail_w])
            centered.setStyle(TableStyle([
                ('ALIGN',   (0,0), (-1,-1), 'CENTER'),
                ('VALIGN',  (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING',  (0,0), (-1,-1), 0),
                ('RIGHTPADDING', (0,0), (-1,-1), 0),
                ('TOPPADDING',   (0,0), (-1,-1), 0),
                ('BOTTOMPADDING',(0,0), (-1,-1), 0),
            ]))
            story.append(centered)
        else:
            story.append(grid_table)

        # ════════════════════════════════════════════════════
        # 2페이지: 단서
        # ════════════════════════════════════════════════════
        story.append(PageBreak())
        story.append(Paragraph("단서 (Clues)", page_title_style))
        story.append(Paragraph(info_text, sub_style))
        story.append(Spacer(1, 4*mm))

        def make_clue_list(word_list, symbol):
            items = [Paragraph(symbol, heading_style)]
            for p in word_list:
                num  = p.get('number', '?')
                txt  = f"{num}.  {p['clue']}  ({len(p['word'])}글자)"
                items.append(Paragraph(txt, clue_style))
            return items

        across_items = make_clue_list(across, "→ 가로")
        down_items   = make_clue_list(down,   "↓ 세로")

        half_w = (avail_w) / 2 - 4*mm
        col_a = KeepInFrame(half_w, 230*mm, across_items, mode='shrink')
        col_d = KeepInFrame(half_w, 230*mm, down_items,   mode='shrink')
        clue_table = Table([[col_a, col_d]], colWidths=[half_w + 4*mm, half_w + 4*mm])
        clue_table.setStyle(TableStyle([
            ('VALIGN',          (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING',     (0,0), (-1,-1), 6),
            ('RIGHTPADDING',    (0,0), (-1,-1), 6),
            ('TOPPADDING',      (0,0), (-1,-1), 0),
            ('BOTTOMPADDING',   (0,0), (-1,-1), 0),
            ('LINEAFTER',       (0,0), (0,-1), 0.5, colors.HexColor('#cccccc')),
        ]))
        story.append(clue_table)

        # ════════════════════════════════════════════════════
        # 3페이지: 정답
        # ════════════════════════════════════════════════════
        story.append(PageBreak())
        story.append(Paragraph("정답 (Answer Key)", page_title_style))
        story.append(Paragraph(info_text, sub_style))
        story.append(Spacer(1, 4*mm))

        # 정답 그리드 (글자 표시)
        ans_num_style = ParagraphStyle(
            'kor_ans_num', fontName=KOR_FONT, fontSize=5,
            textColor=colors.HexColor('#003399'), leading=6,
        )
        ans_letter_style = ParagraphStyle(
            'kor_ans_letter', fontName=KOR_FONT_BOLD, fontSize=int(CELL_MM * 0.55),
            textColor=colors.HexColor('#0f3460'), alignment=TA_CENTER, leading=int(CELL_MM * 0.7),
        )

        # 정답 그리드 — 각 셀에 번호+글자 함께 표시
        ans_table_data = []
        for r in range(rows):
            row_data = []
            for c in range(cols):
                val = self.current_grid[r][c]
                if val == '#':
                    row_data.append('')
                else:
                    num_str = str(self.numbered[(r, c)]) if (r, c) in self.numbered else ''
                    # 번호와 글자를 함께 담은 단락 리스트
                    cell_content = []
                    if num_str:
                        cell_content.append(Paragraph(num_str, ans_num_style))
                    cell_content.append(Paragraph(val, ans_letter_style))
                    row_data.append(cell_content)
            ans_table_data.append(row_data)

        ans_grid = Table(ans_table_data, colWidths=col_widths, rowHeights=row_heights)
        ans_ts = TableStyle([
            ('GRID',            (0,0), (-1,-1), 0.7, colors.HexColor('#2d4a7a')),
            ('VALIGN',          (0,0), (-1,-1), 'TOP'),
            ('ALIGN',           (0,0), (-1,-1), 'CENTER'),
            ('LEFTPADDING',     (0,0), (-1,-1), 1),
            ('TOPPADDING',      (0,0), (-1,-1), 1),
            ('RIGHTPADDING',    (0,0), (-1,-1), 1),
            ('BOTTOMPADDING',   (0,0), (-1,-1), 0),
        ])
        for r in range(rows):
            for c in range(cols):
                if self.current_grid[r][c] == '#':
                    ans_ts.add('BACKGROUND', (c,r), (c,r), colors.HexColor('#1a1a2e'))
                    ans_ts.add('GRID',       (c,r), (c,r), 0, colors.HexColor('#1a1a2e'))
                else:
                    ans_ts.add('BACKGROUND', (c,r), (c,r), colors.HexColor('#eef4ff'))
        ans_grid.setStyle(ans_ts)

        if grid_offset > 0:
            centered_ans = Table([[ans_grid]], colWidths=[avail_w])
            centered_ans.setStyle(TableStyle([
                ('ALIGN',           (0,0), (-1,-1), 'CENTER'),
                ('VALIGN',          (0,0), (-1,-1), 'TOP'),
                ('LEFTPADDING',     (0,0), (-1,-1), 0),
                ('RIGHTPADDING',    (0,0), (-1,-1), 0),
                ('TOPPADDING',      (0,0), (-1,-1), 0),
                ('BOTTOMPADDING',   (0,0), (-1,-1), 0),
            ]))
            story.append(centered_ans)
        else:
            story.append(ans_grid)

        # ════════════════════════════════════════════════════
        # 4페이지: 정답 단어 목록
        # ════════════════════════════════════════════════════
        story.append(PageBreak())
        story.append(Paragraph("정답 단어 목록 (Answer List)", page_title_style))
        story.append(Paragraph(info_text, sub_style))
        story.append(Spacer(1, 4*mm))

        # ── 4페이지 전용 중앙 정렬 스타일 추가 ──
        center_clue_style = ParagraphStyle(
            'kor_center_clue',
            parent=clue_style,
            alignment=TA_CENTER,
            leftIndent=0,  # 기존 들여쓰기 제거 필수
        )

        ans_rows = []
        # 번호 오름차순, 같은 번호는 가로(A) → 세로(D) 순
        for p in sorted(self.current_placed,
                        key=lambda x: (x.get('number', 99), 0 if x['direction'] == 'A' else 1)):
            d_sym = '→' if p['direction'] == 'A' else '↓'
            ans_rows.append([
                Paragraph(f"{p.get('number','?')}", center_clue_style), # 중앙 정렬 스타일 적용
                Paragraph(d_sym, center_clue_style),                    # 중앙 정렬 스타일 적용
                Paragraph(p['clue'], clue_style),
                Paragraph(p['word'], answer_style),
            ])
        if ans_rows:
            # 1열 폭을 12mm로 줄여 1~3자리 모두 자연스럽게 배치 (총합 168mm로 A4 가용폭 내에 안정적으로 안착)
            ans_list_table = Table(ans_rows, colWidths=[12*mm, 12*mm, 104*mm, 40*mm])
            ans_list_table.setStyle(TableStyle([
                ('FONTNAME',        (0,0), (-1,-1), KOR_FONT),
                ('FONTSIZE',        (0,0), (-1,-1), 10),
                ('ROWBACKGROUNDS',  (0,0), (-1,-1), [colors.HexColor('#eef4ff'), colors.white]),
                ('GRID',            (0,0), (-1,-1), 0.3, colors.HexColor('#cccccc')),
                ('VALIGN',          (0,0), (-1,-1), 'MIDDLE'),
                # 좌우 패딩을 대칭(4)으로 맞추어 완벽한 텍스트 중앙 정렬 구현
                ('LEFTPADDING',     (0,0), (-1,-1), 4),
                ('RIGHTPADDING',    (0,0), (-1,-1), 4),
                ('TOPPADDING',      (0,0), (-1,-1), 5),
                ('BOTTOMPADDING',   (0,0), (-1,-1), 5),
            ]))
            story.append(ans_list_table)

        doc.build(story)


# ════════════════════════════════════════════════════════════
# 실행
# ════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = CrosswordApp()
    app.mainloop()
